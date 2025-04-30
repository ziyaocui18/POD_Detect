import os
import re
import time
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook

# df = pd.read_excel("Dillard Street Durham NC/DURHAM NC - Mid June.xlsx", skiprows=8, header=0)
# df = df[["CUSTOMER PO#", "OI#", "SKU", "PRICE", "TALLY", "CHECK POD"]]
# df = df[df["OI#"].notna()]
# print(df)

# List of folders to monitor (each should contain "orders.xlsx" and a "PODs" subfolder)
base_dirs = ['Dillard Street Durham NC']

# Acceptable POD file extensions
valid_extensions = ('.pdf', '.jpg', '.jpeg', '.png', '.docx')

class PODHandler(FileSystemEventHandler):
    """
    Handles POD file detection and Excel updates.
    """
    def __init__(self, folder):
        self.folder = folder
        self.pod_path = os.path.join(folder, "PODs")
        self.excel_path = os.path.join(folder, "DURHAM NC - Mid June.xlsx")
        print(f"\n📂 Watching: {self.pod_path}")
        self.process_existing_pods()

    def extract_order_number(self, filename):
        """
        Extracts the first 5-digit number from a filename.
        """
        match = re.search(r'\b(\d{5})\b', filename)
        return match.group(1) if match else None

    def update_excel_with_order(self, order_number, source_type="new"):
        """
        Checks the Excel file for a matching OI# in column P and updates column V if matched.
        """
        try:
            wb = load_workbook(self.excel_path)
            ws = wb.active

            matched = False
            for row in ws.iter_rows(min_row=9):  # Start from row 9
                cell_value = str(row[15].value).strip() if row[15].value else ""  # Column P (index 15)
                if cell_value == order_number:
                    row[21].value = "✓"  # Column V (index 21)
                    matched = True
                    print(f"✅ {'[EXISTING]' if source_type == 'existing' else '[NEW]'} OI# {order_number} matched → ✓ marked in Column V")
                    break

            if matched:
                wb.save(self.excel_path)
            else:
                print(f"❌ {'[EXISTING]' if source_type == 'existing' else '[NEW]'} OI# {order_number} NOT found in Excel: {self.excel_path}")

        except Exception as e:
            print(f"⚠️ Error updating Excel for OI# {order_number}: {e}")

    def process_existing_pods(self):
        """
        Scan and process already existing POD files at startup.
        """
        print(f"🔎 Scanning existing POD files in {self.pod_path}")
        for fname in os.listdir(self.pod_path):
            if fname.lower().endswith(valid_extensions):
                order_number = self.extract_order_number(fname)
                if order_number:
                    print(f"📄 Found existing POD file: {fname} → OI#: {order_number}")
                    self.update_excel_with_order(order_number, source_type="existing")
                else:
                    print(f"⚠️ Skipped (no valid OI# found): {fname}")

    def on_created(self, event):
        """
        Called when a new file is added to the PODs folder.
        """
        if event.is_directory or not event.src_path.lower().endswith(valid_extensions):
            return

        filename = os.path.basename(event.src_path)
        order_number = self.extract_order_number(filename)
        if order_number:
            print(f"🆕 New POD file detected: {filename} → OI#: {order_number}")
            self.update_excel_with_order(order_number, source_type="new")
        else:
            print(f"⚠️ Skipped new file (no valid OI# found): {filename}")

# Set up and start observers
observers = []

for folder in base_dirs:
    pod_folder = os.path.join(folder, "PODs")
    handler = PODHandler(folder)
    observer = Observer()
    observer.schedule(handler, path=pod_folder, recursive=False)
    observer.start()
    observers.append(observer)

try:
    print("\n👀 Watching for new PODs... Press Ctrl+C to stop.\n")
    while True:
        time.sleep(1)
except KeyboardInterrupt:
    print("\n🛑 Stopping watchers...")
    for observer in observers:
        observer.stop()
    for observer in observers:
        observer.join()
    print("✅ All watchers stopped.")
